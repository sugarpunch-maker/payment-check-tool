import json
import os
import streamlit as st
import pandas as pd
import tempfile
import smtplib
from email.message import EmailMessage

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "mail_config.json")

# ===============================
# UIカラー（コントラスト改善）※修正
# ===============================
st.markdown("""
<style>
.stApp {
    background-color: #DCEEFF;  /* 薄い水色 */
}
input, .stTextInput>div>div>input {
    background-color: #F5F7FA !important;  /* 少しだけグレー */
    color: #333333 !important;
}
.stNumberInput input {
    background-color: #F5F7FA !important;
    color: #333333 !important;
}
.stButton>button {
    background-color: #5DADE2;  /* 水色系ボタン */
    color: white;
}
</style>
""", unsafe_allow_html=True)

# ===============================
# 初回メール保存 → 削除し、新方式へ（修正）
# ===============================
if "smtp_user" not in st.session_state:
    st.session_state["smtp_user"] = ""
    st.session_state["smtp_pass"] = ""

# ===============================
# データ整形
# ===============================
def clean_data(df_orders, df_paid):
    df_orders.columns = df_orders.columns.str.strip()
    df_paid.columns = df_paid.columns.str.strip()
    return df_orders, df_paid

# ===============================
# 組み合わせ計算
# ===============================
def find_best_combination(records, target, max_items=20):

    best_combo = []
    best_sum = 0

    def dfs(start, current, total):
        nonlocal best_combo, best_sum

        # 件数制限
        if len(current) > max_items:
            return

        # target以下で最大更新
        if total <= target and total > best_sum:
            best_combo = current[:]
            best_sum = total

        # 超えたら打ち切り
        if total > target:
            return

        for i in range(start, len(records)):
            dfs(
                i + 1,
                current + [records[i]],
                total + records[i][1]
            )

    dfs(0, [], 0)

    return best_combo

# ===============================
# Excel生成（ローカル保存しない）
# ===============================
def create_excel(df_orders, best_pono):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    file_path = tmp.name
    tmp.close()

    df_orders.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        pono = ws.cell(row=row, column=8).value
        if pono in best_pono:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(file_path)
    return file_path

# ===============================
# メール送信（修正：ユーザー入力使用）
# ===============================
def send_mail(file_path, recipient, smtp_user, smtp_pass):
    try:
        msg = EmailMessage()
        msg["Subject"] = "支払チェック結果"
        msg["From"] = smtp_user
        msg["To"] = recipient
        msg.set_content("支払チェック結果を送付します。")

        with open(file_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename="result.xlsx"
            )

        with smtplib.SMTP("smtp.sina.net", 587) as smtp:
            smtp.starttls()
            smtp.login(smtp_user, smtp_pass)
            smtp.send_message(msg)

        return True, "送信成功"

    except Exception as e:
        return False, str(e)

# ===============================
# メール設定読み込み
# ===============================
if "smtp_user" not in st.session_state:
    st.session_state["smtp_user"] = ""
    st.session_state["smtp_pass"] = ""

if os.path.exists(CONFIG_FILE):
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
        st.session_state["smtp_user"] = config.get("smtp_user", "")
        st.session_state["smtp_pass"] = config.get("smtp_pass", "")
        
# ===============================
# UI
# ===============================
st.title("支払チェックツール")

file_orders = st.file_uploader("注文書整理ファイル")
file_paid = st.file_uploader("送金入金一覧ファイル")

# 修正：初期値0
target = st.number_input("目標金額", value=0)

# ===============================
# メール設定（初回のみ）※修正
# ===============================
if st.session_state["smtp_user"] == "":
    st.subheader("メール設定（初回のみ）")

    smtp_user_input = st.text_input("送信元メールアドレス")
    smtp_pass_input = st.text_input("メールパスワード", type="password")

    if st.button("メール設定を保存"):
        st.session_state["smtp_user"] = smtp_user_input
        st.session_state["smtp_pass"] = smtp_pass_input

        # JSON保存
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({
                    "smtp_user": smtp_user_input,
                    "smtp_pass": smtp_pass_input
                }, f)
        except:
            st.error("設定ファイル保存エラー")

        st.success("保存しました")


else:
    st.write(f"送信元: {st.session_state['smtp_user']}")

recipient = st.text_input("送信先メールアドレス")

# ===============================
# 計算
# ===============================
if st.button("計算", key="calc_button"):

    if file_orders is None or file_paid is None:
        st.error("ファイル不足")
        st.stop()

    df_orders = pd.read_excel(file_orders, header=1)
    df_paid = pd.read_excel(file_paid, header=1)

    df_orders, df_paid = clean_data(df_orders, df_paid)

    col_status = "支払状況"
    col_pono   = "PONO"
    col_amount = "金額"


    df_orders[col_status] = df_orders[col_status].astype(str).str.strip()

    # 支払状況でフィルタ
    df_unpaid = df_orders[
        ~(df_orders["支払状況"].astype(str).str.contains("支払済", na=False))
        ]

    # 右側だけ抽出
    status_idx = df_orders.columns.get_loc("支払状況")
    df_target = df_unpaid.iloc[:, status_idx+1:]

    # 使用列
    col_pono   = "PONO.1"
    col_amount = "金額.1"

    records = []
    
    for _, row in df_target.iterrows():
        try:
            pono = str(row[col_pono]).strip()
            # ★ PONOが空ならスキップ
            if pono == "" or pono.lower() == "nan":
                continue

            val = str(row[col_amount])
            val = val.replace(",", "").replace("円", "").replace("¥", "").strip()
            amount = float(val)

            records.append((pono, amount))
        except:
            continue


    records = sorted(records, key=lambda x: x[1], reverse=True)[:20]

    best_combo = find_best_combination(records, target, 20)
    if not best_combo:
        st.warning("条件内の組み合わせが見つかりませんでした")

    st.session_state["result"] = best_combo
    st.session_state["df_orders"] = df_orders

    st.success("計算完了")

# ===============================
# 結果表示
# ===============================
if "result" in st.session_state:

    st.subheader("計算結果")
    result_df = pd.DataFrame(st.session_state["result"], columns=["PONO", "金額"])
    st.dataframe(result_df)

    if st.button("Excel生成", key="create_excel"):

        best_pono = [x[0] for x in st.session_state["result"]]
        df_orders = st.session_state["df_orders"]

        file_path = create_excel(df_orders, best_pono)

        with open(file_path, "rb") as f:
            st.download_button(
                label="Excelダウンロード",
                data=f,
                file_name="支払チェック結果.xlsx"
            )

        st.session_state["file_path"] = file_path

    # 修正：メール送信
    if st.button("メール送信", key="send_button"):

        if "file_path" not in st.session_state:
            st.error("先にExcel生成してください")

        elif st.session_state["smtp_user"] == "":
            st.error("先にメール設定を登録してください")

        else:
            success, message = send_mail(
                st.session_state["file_path"],
                recipient,
                st.session_state["smtp_user"],
                st.session_state["smtp_pass"]
            )

            if success:
                st.success("メール送信完了")
            else:
                st.error(message)

    if st.button("再計算", key="recalc_button"):
        st.session_state.clear()
        st.experimental_rerun()
